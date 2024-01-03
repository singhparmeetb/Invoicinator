import pdfplumber
import re
from openpyxl import load_workbook

def parse_invoice_local(pdf_path):
    with pdfplumber.open(pdf_path) as pdf:
        text = ""
        for page in pdf.pages:
            text += page.extract_text()
    return text

def parse_invoice(text):
    invoice_number_pattern = re.compile(r"Invoice Number (\S+)")
    invoice_date_pattern = re.compile(r"Invoice Date (\w+ \d+, \d+)")
    gross_total_pattern = re.compile(r"Sub Total \$([\d.]+)")
    tax_pattern = re.compile(r"Tax \$([\d.]+)")

    # Find matches using regular expressions
    gross_total_match = gross_total_pattern.search(text)
    tax_match = tax_pattern.search(text)

    # Extract information from matches or set to "NA" if not found
    gross_total = float(gross_total_match.group(1)) if gross_total_match else 0
    tax_amount = float(tax_match.group(1)) if tax_match else 0
    net_amount = gross_total - tax_amount

    invoice_number_match = invoice_number_pattern.search(text)
    invoice_date_match = invoice_date_pattern.search(text)

    # Extract information from matches or set to "NA" if not found
    invoice_number = invoice_number_match.group(1) if invoice_number_match else "NA"
    invoice_date = invoice_date_match.group(1) if invoice_date_match else "NA"

    return {
        "Invoice Number": invoice_number,
        "Invoice Date": invoice_date,
        "Gross Total": gross_total,
        "Tax Amount": tax_amount,
        "Net Amount": net_amount,
    }

def parse_invoice_quantity_modified(text):
    # Find the position of the headers as a single line
    headers_line_pattern = re.compile(r"\b(Hrs/Qty)\b(.+)$")
    headers_line_match = headers_line_pattern.search(text)

    if headers_line_match:
        # Get the position of the "Qty" in the headers line
        qty_position = headers_line_match.group(2).find("Qty")

        # Define a pattern for extracting quantities in subsequent lines
        quantity_pattern = re.compile(r"^(.{" + str(qty_position) + r"}\d+\.\d{2})")

        # Find matches using the quantity pattern
        quantities = []
        for line in text.split('\n'):
            match = quantity_pattern.match(line)
            if match:
                quantities.append(int(match.group(1)))
            else:
                break  # Break if the line doesn't contain enough words for a table

        return {"Quantities": quantities}
    else:
        return {"Quantities": []}

# def find_first_non_empty_row(sheet):
#     for row in sheet.iter_rows(min_row=1, max_row=sheet.max_row, max_col=1, values_only=True):
#         if any(cell for cell in row if cell is not None and cell != ''):
#             return row[0]
#     return None

def write_to_excel(data, sheet_name="Sheet1"):
    workbook = load_workbook("invoice_data.xlsx")
    sheet = workbook[sheet_name]

    # Find the first non-empty row
    first_row = sheet.max_row + 1

    # Write data to the respective columns
    sheet.cell(row=first_row, column=1, value=data["Invoice Number"])
    sheet.cell(row=first_row, column=2, value=data["Invoice Date"])
    
    quantities = data["Quantities"]
    ans=0
    for i in enumerate(quantities):
        ans+=i
    sheet.cell(row=first_row, column=3, value=ans)

    sheet.cell(row=first_row, column=4, value=data["Gross Total"])
    sheet.cell(row=first_row, column=5, value=data["Tax Amount"])
    sheet.cell(row=first_row, column=6, value=data["Net Amount"])

    # Save the workbook
    workbook.save("invoice_data.xlsx")

def main():
    pdf_path = "invoice.pdf"
    pdf_text = parse_invoice_local(pdf_path)

    invoice_data = parse_invoice(pdf_text)
    quantity_data = parse_invoice_quantity_modified(pdf_text)

    # Combine the data dictionaries
    data = {**invoice_data, **quantity_data}

    # Write the data to Excel
    write_to_excel(data)

if __name__ == "__main__":
    main()
